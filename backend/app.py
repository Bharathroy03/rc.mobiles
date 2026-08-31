# -*- coding: utf-8 -*-
import os
import sys
import sqlite3
import secrets
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
from dotenv import load_dotenv
from supabase import create_client, Client
from werkzeug.security import generate_password_hash, check_password_hash

# Load environment variables
load_dotenv()

DEFAULT_SUPABASE_URL = "https://bjxozcjubwzvuqanvqsp.supabase.co"
DEFAULT_SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImJqeG96Y2p1Ynd6dnVxYW52cXNwIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4ODA3MjMxMSwiZXhwIjoyMTAzNjQ4MzExfQ.SA0ToZWA1HdGjexNefDz77A2CVpZWv2tPpJ6nuKxT_c"

SUPABASE_URL = os.getenv("SUPABASE_URL", DEFAULT_SUPABASE_URL)
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY", os.getenv("SUPABASE_KEY", DEFAULT_SUPABASE_KEY))

# Initialize Supabase Client
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# Flask Application Setup - dynamically locate frontend folder across environments
base_dir = os.path.dirname(os.path.abspath(__file__))
frontend_dir = os.path.abspath(os.path.join(base_dir, "..", "frontend"))
if not os.path.exists(frontend_dir):
    frontend_dir = os.path.abspath(os.path.join(base_dir, "frontend"))
if not os.path.exists(frontend_dir):
    frontend_dir = os.path.abspath(os.path.join(os.getcwd(), "frontend"))

app = Flask(__name__, static_folder=frontend_dir, static_url_path="")
CORS(app)

@app.route("/")
def serve_root():
    if os.path.exists(os.path.join(app.static_folder, "index.html")):
        return send_from_directory(app.static_folder, "index.html")
    return "RC Mobiles API Server Online"

# Ensure upload directory exists (handle serverless /tmp if read-only filesystem)
IS_SERVERLESS = bool(os.environ.get("VERCEL") or os.environ.get("AWS_LAMBDA_FUNCTION_NAME"))
if IS_SERVERLESS:
    UPLOAD_FOLDER = "/tmp/uploads"
    AUTH_DB_PATH = "/tmp/rc_mobiles.db"
    orig_db = os.path.join(os.path.dirname(os.path.abspath(__file__)), "rc_mobiles.db")
    if os.path.exists(orig_db) and not os.path.exists(AUTH_DB_PATH):
        try:
            import shutil
            shutil.copyfile(orig_db, AUTH_DB_PATH)
        except Exception:
            pass
else:
    UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
    AUTH_DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "rc_mobiles.db")

try:
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
except Exception:
    pass
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

def init_auth_db():
    conn = sqlite3.connect(AUTH_DB_PATH)
    cur = conn.cursor()
    cur.execute("""
    CREATE TABLE IF NOT EXISTS app_users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        email TEXT,
        password_hash TEXT NOT NULL,
        full_name TEXT,
        role TEXT DEFAULT 'staff',
        is_active INTEGER DEFAULT 1,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    """)
    # License keys table
    cur.execute("""
    CREATE TABLE IF NOT EXISTS license_keys (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        license_key TEXT UNIQUE NOT NULL,
        plan_name TEXT DEFAULT 'Advanced',
        is_valid INTEGER DEFAULT 1,
        activated_at TIMESTAMP,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    """)
    defaults = [
        ("admin", "admin@rcmobiles.com", generate_password_hash("admin@123"), "Store Owner / Admin", "admin"),
        ("manager", "manager@rcmobiles.com", generate_password_hash("manager@123"), "Store Manager", "manager"),
        ("sales", "sales@rcmobiles.com", generate_password_hash("sales@123"), "Sales Executive", "staff")
    ]
    for u in defaults:
        cur.execute("""
            INSERT INTO app_users (username, email, password_hash, full_name, role, is_active)
            VALUES (?, ?, ?, ?, ?, 1)
            ON CONFLICT(username) DO NOTHING
        """, u)

    # Pre-seed valid license keys (multi-device, no device binding)
    valid_keys = [
        ("RCMB-ADVC-2024-PRO1", "Advanced Pro"),
        ("RCMB-ADVC-2024-PRO2", "Advanced Pro"),
        ("RCMB-STAF-2024-LIC1", "Advanced Staff"),
        ("RCMB-STAF-2024-LIC2", "Advanced Staff"),
        ("RCMB-PREM-2025-ENT1", "Enterprise"),
        ("RCMB-PREM-2025-ENT2", "Enterprise"),
        ("RCMB-DEMO-2025-TST1", "Advanced Pro"),
        ("RCMB-DEMO-2025-TST2", "Advanced Staff"),
        ("HREM-FULL-2025-ADV1", "Advanced Pro"),
        ("HREM-FULL-2025-ADV2", "Advanced Pro"),
    ]
    for key_val, plan in valid_keys:
        cur.execute("""
            INSERT INTO license_keys (license_key, plan_name, is_valid)
            VALUES (?, ?, 1)
            ON CONFLICT(license_key) DO NOTHING
        """, (key_val, plan))

    try:
        conn.commit()
        conn.close()
    except Exception as e:
        print("Warning: init_auth_db commit error:", e)

try:
    init_auth_db()
except Exception as e:
    print("Warning: init_auth_db startup error:", e)

# ---------------- GLOBAL EXCEPTION & ERROR HANDLERS ----------------
@app.errorhandler(400)
def handle_bad_request(e):
    return jsonify({
        "success": False,
        "error": "Bad Request",
        "message": str(e.description if hasattr(e, 'description') else e),
        "code": 400
    }), 400

@app.errorhandler(404)
def handle_not_found(e):
    if request.path.startswith("/api/"):
        return jsonify({
            "success": False,
            "error": "Not Found",
            "message": f"Endpoint '{request.path}' does not exist.",
            "code": 404
        }), 404
    return send_from_directory(app.static_folder, "index.html")

@app.errorhandler(500)
def handle_server_error(e):
    return jsonify({
        "success": False,
        "error": "Internal Server Error",
        "message": str(e),
        "code": 500
    }), 500

@app.errorhandler(Exception)
def handle_generic_exception(e):
    return jsonify({
        "success": False,
        "error": "Unhandled Server Exception",
        "message": str(e),
        "code": 500
    }), 500


# ---------------- API: LICENSE MANAGEMENT ----------------

@app.route("/api/license/activate", methods=["POST"])
def activate_license():
    """Validate and activate a product license key. Multi-device - no device binding."""
    data = request.json or {}
    raw_key = (data.get("license_key") or "").strip().upper()

    if not raw_key:
        return jsonify({"success": False, "error": "License key is required"}), 400

    # Normalize: allow with or without dashes
    normalized = raw_key.replace(" ", "-").upper()

    try:
        conn = sqlite3.connect(AUTH_DB_PATH)
        cur = conn.cursor()
        cur.execute(
            "SELECT id, license_key, plan_name, is_valid, activated_at FROM license_keys WHERE UPPER(license_key) = ?",
            (normalized,)
        )
        row = cur.fetchone()

        if not row:
            conn.close()
            return jsonify({
                "success": False,
                "error": "Invalid license key. Please check the key and try again, or contact your vendor."
            }), 400

        lic_id, lic_key, plan_name, is_valid, activated_at = row

        if not is_valid:
            conn.close()
            return jsonify({
                "success": False,
                "error": "This license key has been revoked or is no longer valid."
            }), 403

        # Mark key as activated (record first use time, but allow multi-device)
        if not activated_at:
            cur.execute(
                "UPDATE license_keys SET activated_at = ? WHERE id = ?",
                (datetime.now().isoformat(), lic_id)
            )
            conn.commit()

        conn.close()

        return jsonify({
            "success": True,
            "message": f"License activated successfully! Welcome to RC Mobiles {plan_name} Plan.",
            "license_key": lic_key,
            "plan_name": plan_name,
            "activated_at": activated_at or datetime.now().isoformat()
        })

    except Exception as err:
        print("License activate error:", err)
        return jsonify({"success": False, "error": str(err)}), 500


@app.route("/api/license/status", methods=["POST"])
def check_license_status():
    """Check whether a stored license key is still valid."""
    data = request.json or {}
    raw_key = (data.get("license_key") or "").strip().upper()

    if not raw_key:
        return jsonify({"active": False, "error": "No license key provided"}), 400

    try:
        conn = sqlite3.connect(AUTH_DB_PATH)
        cur = conn.cursor()
        cur.execute(
            "SELECT id, plan_name, is_valid, activated_at FROM license_keys WHERE UPPER(license_key) = ?",
            (raw_key,)
        )
        row = cur.fetchone()
        conn.close()

        if not row:
            return jsonify({"active": False, "error": "License key not found"})

        _, plan_name, is_valid, activated_at = row
        return jsonify({
            "active": bool(is_valid),
            "plan_name": plan_name,
            "activated_at": activated_at
        })

    except Exception as err:
        print("License status error:", err)
        return jsonify({"active": False, "error": str(err)}), 500


@app.route("/api/license/keys", methods=["GET"])
def list_license_keys():
    """Admin-only: list all license keys (never exposes raw keys in UI - for debugging)."""
    try:
        conn = sqlite3.connect(AUTH_DB_PATH)
        cur = conn.cursor()
        cur.execute("SELECT id, plan_name, is_valid, activated_at, created_at FROM license_keys ORDER BY id")
        rows = cur.fetchall()
        conn.close()
        keys = [
            {"id": r[0], "plan_name": r[1], "is_valid": bool(r[2]),
             "activated_at": r[3], "created_at": r[4]}
            for r in rows
        ]
        return jsonify({"success": True, "total": len(keys), "keys": keys})
    except Exception as err:
        return jsonify({"success": False, "error": str(err)}), 500


# ---------------- FRONTEND ROUTING ----------------
@app.route("/")
def serve_index():
    return send_from_directory(app.static_folder, "index.html")

@app.route("/<path:path>")
def serve_static(path):
    if os.path.exists(os.path.join(app.static_folder, path)):
        return send_from_directory(app.static_folder, path)
    return send_from_directory(app.static_folder, "index.html")

# ---------------- API: HEALTH CHECK ----------------
@app.route("/api/health", methods=["GET"])
def health_check():
    return jsonify({
        "status": "healthy",
        "store": "RC Mobiles Invoice API",
        "database": "Supabase Cloud PostgreSQL",
        "supabase_url": SUPABASE_URL,
        "time": datetime.utcnow().isoformat()
    })

# ---------------- API: STORE SETTINGS ----------------
@app.route("/api/settings", methods=["GET"])
def get_settings():
    try:
        st_res = supabase.table("store_settings").select("*").limit(1).execute()
        st = st_res.data[0] if (st_res.data and len(st_res.data) > 0) else {}

        # Fetch max invoice counter from existing invoices to prevent duplicate keys
        inv_res = supabase.table("invoices").select("invoice_number").execute()
        existing_numbers = [inv["invoice_number"] for inv in (inv_res.data or []) if inv.get("invoice_number")]
        max_counter = 1000
        for num in existing_numbers:
            parts = num.split("-")
            if len(parts) >= 3 and parts[-1].isdigit():
                val = int(parts[-1])
                if val > max_counter:
                    max_counter = val
        
        real_next = max_counter + 1

        if not st:
            default_s = {
                "store_name": "RC Mobiles",
                "address": "NTR Circle, Madakasira, Ananthapur (Sri Sathya Sai district region), Andhra Pradesh 515301",
                "gstin": "37APVPR6953F1Z1",
                "phone": "+91 98490 12345",
                "email": "rcmobiles.madakasira@gmail.com",
                "terms": "1. Goods once sold will not be taken back or exchanged without valid invoice.\n2. Warranty claims are governed strictly by original manufacturer policy.\n3. Physical damage, liquid damage & unauthorized repairs void warranty.\n4. Subject to Madakasira Jurisdiction.",
                "logo_path": "/api/uploads/logo.png",
                "invoice_prefix": "RCM",
                "invoice_counter": real_next
            }
            res_ins = supabase.table("store_settings").insert(default_s).execute()
            return jsonify(res_ins.data[0] if res_ins.data else default_s)

        st["invoice_counter"] = real_next
        # Alias: frontend uses both logo_url and logo_path
        if "logo_path" in st and "logo_url" not in st:
            st["logo_url"] = st["logo_path"]
        return jsonify(st)
    except Exception as err:
        print("Get settings error:", err)
        return jsonify({"error": str(err)}), 500

@app.route("/api/settings", methods=["PUT"])
def update_settings():
    data = request.json or {}
    try:
        res = supabase.table("store_settings").select("id").limit(1).execute()
        if res.data and len(res.data) > 0:
            sid = res.data[0]["id"]
            upd_res = supabase.table("store_settings").update(data).eq("id", sid).execute()
            return jsonify({"message": "Settings updated", "settings": upd_res.data[0] if upd_res.data else data})
        else:
            ins_res = supabase.table("store_settings").insert(data).execute()
            return jsonify({"message": "Settings created", "settings": ins_res.data[0] if ins_res.data else data})
    except Exception as err:
        print("Update settings error:", err)
        return jsonify({"error": str(err)}), 500

# ---------------- SECURITY & RATE LIMITING ----------------
login_attempts = {}  # { ip_or_username: {"count": int, "locked_until": timestamp} }

def is_rate_limited(key):
    now = datetime.now()
    record = login_attempts.get(key)
    if record:
        if record.get("locked_until") and now < record["locked_until"]:
            return True, int((record["locked_until"] - now).total_seconds())
        if record.get("locked_until") and now >= record["locked_until"]:
            login_attempts[key] = {"count": 0, "locked_until": None}
    return False, 0

def record_failed_login(key):
    now = datetime.now()
    record = login_attempts.get(key, {"count": 0, "locked_until": None})
    record["count"] += 1
    if record["count"] >= 5:
        from datetime import timedelta
        record["locked_until"] = now + timedelta(minutes=5)
    login_attempts[key] = record

def reset_failed_login(key):
    if key in login_attempts:
        login_attempts[key] = {"count": 0, "locked_until": None}

# ---------------- API: AUTHENTICATION ----------------
@app.route("/api/auth/login", methods=["POST"])
def auth_login():
    data = request.json or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")
    client_ip = request.remote_addr or "unknown"
    rate_key = f"{client_ip}:{username.lower()}"

    # Rate limiting check
    limited, secs_left = is_rate_limited(rate_key)
    if limited:
        return jsonify({
            "success": False,
            "error": f"Too many failed login attempts. Please wait {secs_left} seconds before trying again."
        }), 429

    if not username or not password:
        return jsonify({"success": False, "error": "Username and password are required"}), 400

    user_found = None
    source = "sqlite"

    # 1. Check Supabase users table first if available
    try:
        sb_res = supabase.table("users").select("*").ilike("username", username).eq("is_active", True).limit(1).execute()
        if sb_res.data and len(sb_res.data) > 0:
            user_found = sb_res.data[0]
            source = "supabase"
    except Exception:
        pass

    # 2. Check local SQLite DB if not found in Supabase
    if not user_found:
        try:
            conn = sqlite3.connect(AUTH_DB_PATH)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute("SELECT * FROM app_users WHERE LOWER(username) = LOWER(?) AND is_active = 1", (username,))
            row = cur.fetchone()
            conn.close()
            if row:
                user_found = dict(row)
                source = "sqlite"
        except Exception as err:
            print("SQLite user query error:", err)

    if not user_found:
        record_failed_login(rate_key)
        return jsonify({"success": False, "error": "Invalid username or password"}), 401

    # Verify password (support secure hash, with auto-upgrade if manually entered plain text in SQL)
    pwd_hash = user_found.get("password_hash") or ""
    is_valid = False
    
    if pwd_hash.startswith("scrypt:") or pwd_hash.startswith("pbkdf2:"):
        is_valid = check_password_hash(pwd_hash, password)
    elif pwd_hash == password:  # Manual entry in database
        is_valid = True
        # Upgrade plaintext to hash
        new_hash = generate_password_hash(password)
        if source == "sqlite":
            try:
                conn = sqlite3.connect(AUTH_DB_PATH)
                conn.execute("UPDATE app_users SET password_hash = ? WHERE id = ?", (new_hash, user_found["id"]))
                conn.commit()
                conn.close()
            except Exception:
                pass
        elif source == "supabase":
            try:
                supabase.table("users").update({"password_hash": new_hash}).eq("id", user_found["id"]).execute()
            except Exception:
                pass

    if not is_valid:
        record_failed_login(rate_key)
        return jsonify({"success": False, "error": "Invalid username or password"}), 401

    # Successful login: reset rate limit counter
    reset_failed_login(rate_key)

    token = secrets.token_hex(32)
    user_info = {
        "id": user_found["id"],
        "username": user_found["username"],
        "full_name": user_found.get("full_name") or user_found["username"],
        "email": user_found.get("email") or "",
        "role": (user_found.get("role") or "staff").lower(),
        "token": token
    }

    return jsonify({"success": True, "message": "Authentication successful", "user": user_info})

@app.route("/api/auth/me", methods=["GET"])
def auth_me():
    return jsonify({"success": True, "status": "authenticated"})

@app.route("/api/auth/logout", methods=["POST"])
def auth_logout():
    return jsonify({"success": True, "message": "Logged out successfully"})

# ---------------- API: USER MANAGEMENT (ADMIN / DESIGNATIONS) ----------------
@app.route("/api/users", methods=["GET"])
def get_all_users():
    try:
        # Check Supabase first
        try:
            sb_res = supabase.table("users").select("id, username, email, full_name, role, is_active, created_at").order("id").execute()
            if sb_res.data and len(sb_res.data) > 0:
                return jsonify(sb_res.data)
        except Exception:
            pass

        conn = sqlite3.connect(AUTH_DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT id, username, email, full_name, role, is_active, created_at FROM app_users ORDER BY id")
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return jsonify(rows)
    except Exception as err:
        print("Get users error:", err)
        return jsonify([]), 500

@app.route("/api/users", methods=["POST"])
def create_user():
    data = request.json or {}
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    role = (data.get("role") or "staff").strip().lower()
    full_name = data.get("full_name", "").strip() or username
    email = data.get("email", "").strip()

    if not username or not password:
        return jsonify({"success": False, "error": "Username and password are required"}), 400

    hashed = generate_password_hash(password)

    # Save to SQLite
    new_id = None
    try:
        conn = sqlite3.connect(AUTH_DB_PATH)
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO app_users (username, email, password_hash, full_name, role, is_active)
            VALUES (?, ?, ?, ?, ?, 1)
        """, (username, email, hashed, full_name, role))
        new_id = cur.lastrowid
        conn.commit()
        conn.close()
    except sqlite3.IntegrityError:
        return jsonify({"success": False, "error": "Username already exists"}), 400
    except Exception as err:
        print("Create sqlite user error:", err)

    # Sync to Supabase if table exists
    try:
        supabase.table("users").insert({
            "username": username,
            "email": email,
            "password_hash": hashed,
            "full_name": full_name,
            "role": role,
            "is_active": True
        }).execute()
    except Exception as err:
        print("Supabase user sync notice:", err)

    return jsonify({
        "success": True,
        "message": f"User {username} created successfully",
        "user": {"id": new_id or 1, "username": username, "full_name": full_name, "role": role, "email": email}
    }), 201

@app.route("/api/users/<int:user_id>", methods=["PUT"])
def update_user(user_id):
    data = request.json or {}
    role = data.get("role")
    full_name = data.get("full_name")
    password = data.get("password")
    is_active = data.get("is_active")

    try:
        conn = sqlite3.connect(AUTH_DB_PATH)
        cur = conn.cursor()
        if password:
            cur.execute("""
                UPDATE app_users SET full_name = COALESCE(?, full_name), role = COALESCE(?, role),
                password_hash = ?, is_active = COALESCE(?, is_active) WHERE id = ?
            """, (full_name, role, generate_password_hash(password), is_active, user_id))
        else:
            cur.execute("""
                UPDATE app_users SET full_name = COALESCE(?, full_name), role = COALESCE(?, role),
                is_active = COALESCE(?, is_active) WHERE id = ?
            """, (full_name, role, is_active, user_id))
        conn.commit()
        conn.close()
        return jsonify({"success": True, "message": "User updated successfully"})
    except Exception as err:
        print("Update user error:", err)
        return jsonify({"success": False, "error": str(err)}), 500

@app.route("/api/users/<int:user_id>", methods=["DELETE"])
def delete_user(user_id):
    if user_id == 1:
        return jsonify({"success": False, "error": "Cannot delete primary admin account"}), 400
    try:
        conn = sqlite3.connect(AUTH_DB_PATH)
        cur = conn.cursor()
        cur.execute("DELETE FROM app_users WHERE id = ?", (user_id,))
        conn.commit()
        conn.close()

        try:
            supabase.table("users").delete().eq("id", user_id).execute()
        except Exception:
            pass

        return jsonify({"success": True, "message": "User deleted successfully"})
    except Exception as err:
        print("Delete user error:", err)
        return jsonify({"success": False, "error": str(err)}), 500

# ---------------- API: PRODUCTS CATALOG ----------------
@app.route("/api/products", methods=["GET"])
def get_products():
    query_str = request.args.get("query", "").strip()
    category = request.args.get("category", "").strip()
    
    try:
        req = supabase.table("products").select("*").order("name")
        if query_str:
            # Use individual filters OR-chained (compatible with all supabase-py versions)
            name_res = supabase.table("products").select("*").ilike("name", f"%{query_str}%").order("name").execute()
            brand_res = supabase.table("products").select("*").ilike("brand", f"%{query_str}%").order("name").execute()
            seen_ids = set()
            combined = []
            for row in (name_res.data or []) + (brand_res.data or []):
                if row.get("id") not in seen_ids:
                    seen_ids.add(row["id"])
                    if not category or row.get("category") == category:
                        combined.append(row)
            return jsonify(combined)
        if category:
            req = req.eq("category", category)
        res = req.execute()
        return jsonify(res.data or [])
    except Exception as err:
        print("Get products error:", err)
        return jsonify([]), 500

@app.route("/api/products", methods=["POST"])
def add_product():
    data = request.json or {}
    # Accept both 'price' and 'selling_price' field names
    selling_price = data.get("selling_price") or data.get("price")
    if not data.get("name") or selling_price is None:
        return jsonify({"error": "Product name and selling price are required"}), 400
        
    payload = {
        "name": data["name"],
        "brand": data.get("brand", "Generic"),
        "category": data.get("category", "Mobile"),
        "hsn_code": data.get("hsn_code", "8517"),
        "purchase_price": float(data.get("purchase_price", 0.0)),
        "selling_price": float(selling_price),
        "stock_qty": int(data.get("stock_qty", 1)),
        "tax_rate": float(data.get("tax_rate", 18.0))
    }
    
    try:
        res = supabase.table("products").insert(payload).execute()
        return jsonify({"success": True, "message": "Product added successfully", "id": res.data[0]["id"] if res.data else None, "product": res.data[0] if res.data else payload}), 201
    except Exception as err:
        print("Add product error:", err)
        return jsonify({"success": False, "error": str(err)}), 500

@app.route("/api/products/<int:prod_id>", methods=["PUT"])
def update_product(prod_id):
    data = request.json or {}
    try:
        res = supabase.table("products").update(data).eq("id", prod_id).execute()
        if not res.data or len(res.data) == 0:
            return jsonify({"success": False, "error": "Product not found"}), 404
        return jsonify({"success": True, "message": "Product updated successfully", "product": res.data[0]})
    except Exception as err:
        print("Update product error:", err)
        return jsonify({"success": False, "error": str(err)}), 500

@app.route("/api/products/<int:prod_id>", methods=["DELETE"])
def delete_product(prod_id):
    try:
        res = supabase.table("products").delete().eq("id", prod_id).execute()
        try:
            Product.query.filter_by(id=prod_id).delete()
            db.session.commit()
        except Exception:
            pass
        return jsonify({"success": True, "message": "Product deleted successfully"})
    except Exception as err:
        print("Delete product error:", err)
        return jsonify({"success": False, "error": str(err)}), 500

@app.route("/api/products/brands", methods=["GET"])
def get_product_brands():
    category = request.args.get("category", "").strip()
    try:
        req = supabase.table("products").select("brand, category, stock_qty")
        if category:
            req = req.eq("category", category)
        res = req.execute()
        rows = res.data or []
        
        brand_stats = {}
        for r in rows:
            b = (r.get("brand") or "Generic").strip()
            if not b:
                b = "Generic"
            if b not in brand_stats:
                brand_stats[b] = {"brand": b, "product_count": 0, "total_stock": 0}
            brand_stats[b]["product_count"] += 1
            brand_stats[b]["total_stock"] += int(r.get("stock_qty") or 0)
            
        brands_list = sorted(list(brand_stats.values()), key=lambda x: x["brand"].lower())
        return jsonify(brands_list)
    except Exception as err:
        print("Get product brands error:", err)
        return jsonify([]), 500

@app.route("/api/products/bulk-delete", methods=["POST"])
def bulk_delete_products():
    data = request.json or {}
    product_ids = data.get("product_ids", [])
    if not product_ids:
        return jsonify({"success": False, "error": "No product IDs provided"}), 400
    try:
        res = supabase.table("products").delete().in_("id", product_ids).execute()
        try:
            Product.query.filter(Product.id.in_(product_ids)).delete(synchronize_session=False)
            db.session.commit()
        except Exception:
            pass
        return jsonify({"success": True, "message": f"Successfully deleted {len(product_ids)} products"})
    except Exception as err:
        print("Bulk delete products error:", err)
        return jsonify({"success": False, "error": str(err)}), 500

@app.route("/api/products/by-brand", methods=["DELETE", "POST"])
def delete_products_by_brand():
    data = request.json or {}
    brand = data.get("brand") or request.args.get("brand", "").strip()
    category = data.get("category") or request.args.get("category", "").strip()
    
    if not brand:
        return jsonify({"success": False, "error": "Brand name is required"}), 400
        
    try:
        req = supabase.table("products").delete().ilike("brand", brand)
        if category:
            req = req.eq("category", category)
        res = req.execute()
        try:
            q = Product.query.filter(Product.brand.ilike(brand))
            if category:
                q = q.filter(Product.category == category)
            q.delete(synchronize_session=False)
            db.session.commit()
        except Exception:
            pass
        return jsonify({"success": True, "message": f"Successfully deleted all products under brand '{brand}'"})
    except Exception as err:
        print("Delete products by brand error:", err)
        return jsonify({"success": False, "error": str(err)}), 500

@app.route("/api/products/all", methods=["DELETE", "POST"])
def delete_all_products():
    category = request.args.get("category", "").strip() or (request.json or {}).get("category", "").strip()
    try:
        if category:
            res = supabase.table("products").delete().eq("category", category).execute()
            try:
                Product.query.filter(Product.category == category).delete(synchronize_session=False)
                db.session.commit()
            except Exception:
                pass
            return jsonify({"success": True, "message": f"Successfully deleted all {category} products from inventory"})
        else:
            res = supabase.table("products").delete().neq("id", 0).execute()
            try:
                Product.query.delete()
                db.session.commit()
            except Exception:
                pass
            return jsonify({"success": True, "message": "Successfully wiped all inventory catalog products"})
    except Exception as err:
        print("Delete all products error:", err)
        return jsonify({"success": False, "error": str(err)}), 500

# ---------------- API: INVOICES ----------------
@app.route("/api/invoices", methods=["GET"])
def get_invoices():
    search = request.args.get("search", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    
    try:
        req = supabase.table("invoices").select("*, invoice_items(*)").order("id", desc=True)
        if search:
            # Use individual filters OR-chained
            num_res = supabase.table("invoices").select("*, invoice_items(*)").ilike("invoice_number", f"%{search}%").order("id", desc=True).execute()
            name_res = supabase.table("invoices").select("*, invoice_items(*)").ilike("customer_name", f"%{search}%").order("id", desc=True).execute()
            phone_res = supabase.table("invoices").select("*, invoice_items(*)").ilike("customer_phone", f"%{search}%").order("id", desc=True).execute()
            seen_ids = set()
            invoices = []
            for row in (num_res.data or []) + (name_res.data or []) + (phone_res.data or []):
                if row.get("id") not in seen_ids:
                    seen_ids.add(row["id"])
                    invoices.append(row)
        else:
            res = req.execute()
            invoices = res.data or []
            
        # Apply date filters in Python (after fetch)
        if date_from:
            invoices = [i for i in invoices if (i.get("invoice_date") or "") >= date_from]
        if date_to:
            invoices = [i for i in invoices if (i.get("invoice_date") or "") <= date_to + "T23:59:59"]
            
        # Format dates for UI display
        for inv in invoices:
            dt_str = inv.get("invoice_date") or inv.get("created_at") or ""
            if dt_str:
                try:
                    dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
                    inv["formatted_date"] = dt.strftime("%d-%b-%Y %I:%M %p")
                except Exception:
                    inv["formatted_date"] = dt_str
            else:
                inv["formatted_date"] = ""
            inv["items"] = inv.get("invoice_items", [])
            
        return jsonify(invoices)
    except Exception as err:
        print("Get invoices error:", err)
        return jsonify([]), 500

@app.route("/api/invoices/<int:inv_id>", methods=["GET"])
def get_invoice_by_id(inv_id):
    try:
        res = supabase.table("invoices").select("*, invoice_items(*)").eq("id", inv_id).execute()
        if not res.data or len(res.data) == 0:
            return jsonify({"success": False, "error": "Invoice not found"}), 404
            
        inv = res.data[0]
        dt_str = inv.get("invoice_date") or inv.get("created_at") or ""
        if dt_str:
            try:
                dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
                inv["formatted_date"] = dt.strftime("%d-%b-%Y %I:%M %p")
            except Exception:
                inv["formatted_date"] = dt_str
        else:
            inv["formatted_date"] = ""
        inv["items"] = inv.get("invoice_items", [])
        return jsonify(inv)
    except Exception as err:
        print("Get invoice by ID error:", err)
        return jsonify({"success": False, "error": str(err)}), 500

@app.route("/api/invoices/<int:inv_id>", methods=["PUT"])
def update_invoice(inv_id):
    data = request.json or {}
    items_data = data.get("items", [])
    if not items_data or len(items_data) == 0:
        return jsonify({"success": False, "error": "At least one line item is required"}), 400

    try:
        subtotal = 0.0
        processed_items = []
        
        for item in items_data:
            unit_price = float(item.get("unit_price", 0.0))
            qty = max(1, int(item.get("quantity", 1)))
            item_total = unit_price * qty
            subtotal += item_total
            
            tax_rate = float(item.get("tax_rate", 18.0))
            taxable_val = round((item_total / (1 + (tax_rate / 100))), 2)
            gst_val = round(item_total - taxable_val, 2)
            cgst = round(gst_val / 2, 2)
            sgst = round(gst_val - cgst, 2)
            
            processed_items.append({
                "invoice_id": inv_id,
                "product_id": item.get("product_id"),
                "item_name": item.get("item_name", "Item"),
                "hsn_code": item.get("hsn_code", "8517"),
                "imei_serial": item.get("imei_serial", ""),
                "quantity": qty,
                "unit_price": unit_price,
                "tax_rate": tax_rate,
                "taxable_value": taxable_val,
                "cgst_amount": cgst,
                "sgst_amount": sgst,
                "igst_amount": 0.0,
                "total_amount": item_total
            })

        discount = float(data.get("discount_amount", 0.0))
        grand_total = max(0.0, subtotal - discount)
        taxable_amount = round(grand_total / 1.18, 2)
        total_tax = round(grand_total - taxable_amount, 2)
        cgst_amount = round(total_tax / 2, 2)
        sgst_amount = round(total_tax - cgst_amount, 2)

        inv_payload = {
            "customer_name": data.get("customer_name", "Cash Customer"),
            "customer_phone": data.get("customer_phone", ""),
            "customer_address": data.get("customer_address", "Madakasira, AP"),
            "customer_gstin": data.get("customer_gstin", ""),
            "state_type": data.get("state_type", "INTRA_STATE"),
            "subtotal": subtotal,
            "discount_amount": discount,
            "taxable_amount": taxable_amount,
            "cgst_amount": cgst_amount,
            "sgst_amount": sgst_amount,
            "igst_amount": 0.0,
            "total_tax": total_tax,
            "grand_total": grand_total,
            "payment_mode": data.get("payment_mode", "Cash"),
            "payment_status": data.get("payment_status", "Paid"),
            "notes": data.get("notes", "")
        }
        if data.get("invoice_date"):
            inv_payload["invoice_date"] = data["invoice_date"]

        inv_res = supabase.table("invoices").update(inv_payload).eq("id", inv_id).execute()
        if not inv_res.data or len(inv_res.data) == 0:
            return jsonify({"success": False, "error": "Invoice not found"}), 404

        # Refresh line items
        supabase.table("invoice_items").delete().eq("invoice_id", inv_id).execute()
        for p_item in processed_items:
            supabase.table("invoice_items").insert(p_item).execute()

        updated_inv = inv_res.data[0]
        updated_inv["items"] = processed_items
        return jsonify({"success": True, "message": "Invoice updated successfully", "invoice": updated_inv})

    except Exception as err:
        print("Update invoice error:", err)
        return jsonify({"success": False, "error": str(err)}), 500

@app.route("/api/invoices/<int:inv_id>", methods=["DELETE"])
def delete_invoice(inv_id):
    try:
        supabase.table("invoice_items").delete().eq("invoice_id", inv_id).execute()
        supabase.table("invoices").delete().eq("id", inv_id).execute()
        return jsonify({"success": True, "message": "Invoice deleted successfully"})
    except Exception as err:
        print("Delete invoice error:", err)
        return jsonify({"success": False, "error": str(err)}), 500


@app.route("/api/invoices/all", methods=["DELETE", "POST"])
def delete_all_invoices():
    """Permanently delete ALL invoices and their line items. Irreversible."""
    try:
        # Count invoices before deletion (for reporting)
        count_res = supabase.table("invoices").select("id", count="exact").execute()
        total_count = count_res.count if hasattr(count_res, 'count') and count_res.count else len(count_res.data or [])

        if total_count == 0:
            return jsonify({"success": True, "message": "No invoices to delete.", "deleted": 0})

        # Delete all line items first (foreign key child records)
        supabase.table("invoice_items").delete().neq("id", 0).execute()
        # Delete all invoices
        supabase.table("invoices").delete().neq("id", 0).execute()

        return jsonify({
            "success": True,
            "message": f"All {total_count} invoice(s) and their line items have been permanently deleted.",
            "deleted": total_count
        })
    except Exception as err:
        print("Delete all invoices error:", err)
        return jsonify({"success": False, "error": str(err)}), 500

@app.route("/api/invoices", methods=["POST"])
def create_invoice():
    data = request.json or {}
    items_data = data.get("items", [])
    if not items_data:
        return jsonify({"error": "At least one line item is required"}), 400

    # Retrieve store settings
    st_res = supabase.table("store_settings").select("*").limit(1).execute()
    st_data = st_res.data[0] if st_res.data else {}
    
    prefix = st_data.get("invoice_prefix") or "RCM"
    yyyymm = datetime.utcnow().strftime("%Y%m")

    # Fetch existing invoice numbers to prevent duplicate key errors (23505)
    all_invs = supabase.table("invoices").select("invoice_number").execute()
    existing_numbers = set(inv.get("invoice_number") for inv in (all_invs.data or []) if inv.get("invoice_number"))
    
    max_counter = 1000
    for num in existing_numbers:
        parts = num.split("-")
        if len(parts) >= 3 and parts[-1].isdigit():
            val = int(parts[-1])
            if val > max_counter:
                max_counter = val

    next_counter = max_counter + 1
    # Ensure strictly sequential unique invoice number
    req_number = data.get("invoice_number")
    if not req_number or req_number in existing_numbers:
        inv_number = f"{prefix}-{yyyymm}-{next_counter}"
        current_used_counter = next_counter
        future_next_counter = next_counter + 1
    else:
        inv_number = req_number
        current_used_counter = next_counter
        future_next_counter = next_counter + 1

    # Update store settings with future_next_counter
    if st_data.get("id"):
        supabase.table("store_settings").update({"invoice_counter": future_next_counter}).eq("id", st_data["id"]).execute()

    # Calculations
    subtotal = 0.0
    processed_items = []
    
    for item in items_data:
        unit_price = float(item.get("unit_price", 0.0))
        qty = int(item.get("quantity", 1))
        item_total = unit_price * qty
        subtotal += item_total
        
        tax_rate = float(item.get("tax_rate", 18.0))
        taxable_val = round((item_total / (1 + (tax_rate / 100))), 2)
        gst_val = round(item_total - taxable_val, 2)
        cgst = round(gst_val / 2, 2)
        sgst = round(gst_val - cgst, 2)
        
        processed_items.append({
            "product_id": item.get("product_id"),
            "item_name": item.get("item_name", "Item"),
            "hsn_code": item.get("hsn_code", "8517"),
            "imei_serial": item.get("imei_serial", ""),
            "quantity": qty,
            "unit_price": unit_price,
            "tax_rate": tax_rate,
            "taxable_value": taxable_val,
            "cgst_amount": cgst,
            "sgst_amount": sgst,
            "igst_amount": 0.0,
            "total_amount": item_total
        })

    discount = float(data.get("discount_amount", 0.0))
    grand_total = max(0.0, subtotal - discount)
    taxable_amount = round(grand_total / 1.18, 2)
    total_tax = round(grand_total - taxable_amount, 2)
    cgst_amount = round(total_tax / 2, 2)
    sgst_amount = round(total_tax - cgst_amount, 2)

    inv_payload = {
        "invoice_number": inv_number,
        "invoice_date": datetime.utcnow().isoformat(),
        "customer_name": data.get("customer_name", "Cash Customer"),
        "customer_phone": data.get("customer_phone", ""),
        "customer_address": data.get("customer_address", "Madakasira, AP"),
        "customer_gstin": data.get("customer_gstin", ""),
        "state_type": data.get("state_type", "INTRA_STATE"),
        "subtotal": subtotal,
        "discount_amount": discount,
        "taxable_amount": taxable_amount,
        "cgst_amount": cgst_amount,
        "sgst_amount": sgst_amount,
        "igst_amount": 0.0,
        "total_tax": total_tax,
        "grand_total": grand_total,
        "payment_mode": data.get("payment_mode", "Cash"),
        "payment_status": data.get("payment_status", "Paid"),
        "notes": data.get("notes", "")
    }

    try:
        inv_res = supabase.table("invoices").insert(inv_payload).execute()
        if not inv_res.data or len(inv_res.data) == 0:
            return jsonify({"error": "Failed to create invoice in Supabase"}), 500

        new_invoice = inv_res.data[0]
        new_inv_id = new_invoice["id"]

        # Insert line items
        for p_item in processed_items:
            p_item["invoice_id"] = new_inv_id
            supabase.table("invoice_items").insert(p_item).execute()

        new_invoice["items"] = processed_items
        next_invoice_number = f"{prefix}-{yyyymm}-{future_next_counter}"
        return jsonify({
            "message": "Invoice created successfully", 
            "invoice": new_invoice,
            "next_counter": future_next_counter,
            "next_invoice_number": next_invoice_number
        }), 201

    except Exception as err:
        print("Create invoice error:", err)
        return jsonify({"error": str(err)}), 500

# ---------------- API: DASHBOARD ANALYTICS ----------------
@app.route("/api/dashboard/stats", methods=["GET"])
def get_dashboard_stats():
    try:
        inv_res = supabase.table("invoices").select("*").execute()
        invoices = inv_res.data or []

        today_str = datetime.utcnow().strftime("%Y-%m-%d")
        month_str = datetime.utcnow().strftime("%Y-%m")

        today_sales = 0.0
        today_count = 0
        month_sales = 0.0
        month_count = 0
        total_sales = 0.0
        total_count = len(invoices)

        for inv in invoices:
            g_total = float(inv.get("grand_total", 0.0))
            total_sales += g_total
            
            inv_date = inv.get("invoice_date") or inv.get("created_at") or ""
            if inv_date.startswith(today_str):
                today_sales += g_total
                today_count += 1
            if inv_date.startswith(month_str):
                month_sales += g_total
                month_count += 1

        prod_res = supabase.table("products").select("id, stock_qty").execute()
        products = prod_res.data or []
        low_stock_count = sum(1 for p in products if p.get("stock_qty", 0) <= 2)

        return jsonify({
            # Primary keys (used by dashboard UI)
            "total_invoices": total_count,
            "total_revenue": round(total_sales, 2),
            "total_products": len(products),
            "low_stock_count": low_stock_count,
            # Alias keys (legacy / analytics)
            "today_sales": round(today_sales, 2),
            "today_invoice_count": today_count,
            "month_sales": round(month_sales, 2),
            "month_invoice_count": month_count,
            "total_sales": round(total_sales, 2),
            "total_invoice_count": total_count
        })
    except Exception as err:
        print("Dashboard stats error:", err)
        return jsonify({"error": str(err)}), 500

# ---------------- API: FILE UPLOADS & LOGO SERVE ----------------
@app.route("/api/uploads/<path:filename>")
def serve_upload(filename):
    # 1. Check in configured UPLOAD_FOLDER
    file_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    if os.path.exists(file_path):
        return send_from_directory(app.config["UPLOAD_FOLDER"], filename)
    
    # 2. Check in backend/uploads directory
    backend_uploads = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
    if os.path.exists(os.path.join(backend_uploads, filename)):
        return send_from_directory(backend_uploads, filename)

    # 3. Check in Images directory (Main_Logo.png)
    images_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Images")
    if "logo" in filename.lower() and os.path.exists(os.path.join(images_dir, "Main_Logo.png")):
        return send_from_directory(images_dir, "Main_Logo.png")
    if os.path.exists(os.path.join(images_dir, filename)):
        return send_from_directory(images_dir, filename)

    # 4. Check in frontend static folder
    if os.path.exists(os.path.join(app.static_folder, filename)):
        return send_from_directory(app.static_folder, filename)
    if "logo" in filename.lower() and os.path.exists(os.path.join(app.static_folder, "logo.png")):
        return send_from_directory(app.static_folder, "logo.png")

    return jsonify({"error": "File not found"}), 404

# ---------------- API: REPORTS EXPORT (CSV & EXCEL) ----------------
@app.route("/api/reports/export", methods=["GET"])
def export_invoices_csv():
    try:
        inv_res = supabase.table("invoices").select("*, invoice_items(*)").order("id", desc=True).execute()
        invoices = inv_res.data or []

        csv_lines = ["Invoice No,Date,Customer Name,Customer Phone,Customer Address,Payment Mode,Total Items,Subtotal,Tax Amount,Discount,Grand Total"]
        for inv in invoices:
            dt_str = (inv.get("invoice_date") or inv.get("created_at") or "").replace('"', '""')
            c_name = (inv.get("customer_name") or "Cash Customer").replace('"', '""')
            c_phone = (inv.get("customer_phone") or "N/A").replace('"', '""')
            c_addr = (inv.get("customer_address") or "N/A").replace('"', '""')
            p_mode = (inv.get("payment_mode") or "Cash").upper()
            items = inv.get("invoice_items") or []
            item_count = len(items)

            sub = float(inv.get("subtotal") or 0.0)
            tax = float(inv.get("tax_amount") or 0.0)
            disc = float(inv.get("discount_amount") or 0.0)
            g_total = float(inv.get("grand_total") or 0.0)

            line = f'"{inv.get("invoice_number")}","{dt_str}","{c_name}","{c_phone}","{c_addr}","{p_mode}",{item_count},{sub:.2f},{tax:.2f},{disc:.2f},{g_total:.2f}'
            csv_lines.append(line)

        csv_content = "\n".join(csv_lines)
        return csv_content, 200, {
            "Content-Type": "text/csv; charset=utf-8",
            "Content-Disposition": "attachment; filename=RC_Mobiles_Invoices_Report.csv"
        }
    except Exception as err:
        print("CSV export error:", err)
        return jsonify({"error": str(err)}), 500

@app.route("/api/reports/export-excel", methods=["GET"])
def export_invoices_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io

        inv_res = supabase.table("invoices").select("*, invoice_items(*)").order("id", desc=True).execute()
        invoices = inv_res.data or []

        wb = openpyxl.Workbook()
        
        # ---------------- SHEET 1: INVOICES SUMMARY ----------------
        ws1 = wb.active
        ws1.title = "Sales Invoices Summary"
        ws1.views.sheetView[0].showGridLines = True

        # Palettes
        c_navy = "001F3F"
        c_blue = "2563EB"
        c_header_fill = "001F3F"
        c_zebra = "F0F4F8"
        c_border = "CBD5E1"

        font_title = Font(name="Calibri", size=16, bold=True, color="001F3F")
        font_subtitle = Font(name="Calibri", size=10, italic=True, color="475569")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=11, color="0F172A")
        font_bold = Font(name="Calibri", size=11, bold=True, color="0F172A")
        font_total = Font(name="Calibri", size=12, bold=True, color="001F3F")

        fill_header = PatternFill(start_color=c_header_fill, end_color=c_header_fill, fill_type="solid")
        fill_zebra = PatternFill(start_color=c_zebra, end_color=c_zebra, fill_type="solid")
        fill_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

        thin_side = Side(border_style="thin", color=c_border)
        double_bottom = Side(border_style="double", color="001F3F")
        top_thin = Side(border_style="thin", color="001F3F")
        border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        border_total = Border(top=top_thin, bottom=double_bottom, left=thin_side, right=thin_side)

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        # Title Block
        ws1.merge_cells("A1:K1")
        ws1["A1"] = "RC MOBILES & SERVICES - GST SALES INVOICE REGISTER"
        ws1["A1"].font = font_title
        ws1["A1"].alignment = align_left

        ws1.merge_cells("A2:K2")
        ws1["A2"] = "NTR Circle, Madakasira, Ananthapur Dist, AP | GSTIN: 37APVPR6953F1Z1 | Report Generated: " + datetime.now().strftime("%d-%b-%Y %I:%M %p")
        ws1["A2"].font = font_subtitle
        ws1["A2"].alignment = align_left

        headers_s1 = [
            "SL #", "Invoice Number", "Invoice Date", "Customer Name", 
            "Mobile Phone", "Address", "Payment Mode", 
            "Items Count", "Subtotal (₹)", "Discount (₹)", "Grand Total (₹)"
        ]

        # Header Row at row 4
        row_idx = 4
        for col_idx, header_text in enumerate(headers_s1, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=header_text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_cell
        ws1.row_dimensions[row_idx].height = 28

        # Populate Invoices
        start_data_row = 5
        cur_row = start_data_row

        for idx, inv in enumerate(invoices, 1):
            items = inv.get("invoice_items") or []
            dt_raw = inv.get("invoice_date") or inv.get("created_at") or ""
            dt_clean = dt_raw.split(".")[0].replace("T", " ") if dt_raw else "N/A"

            c_name = (inv.get("customer_name") or "Cash Customer").strip()
            c_phone = (inv.get("customer_phone") or "N/A").strip()
            c_addr = (inv.get("customer_address") or "N/A").strip()
            p_mode = (inv.get("payment_mode") or "Cash").upper()
            
            sub = float(inv.get("subtotal") or 0.0)
            disc = float(inv.get("discount_amount") or 0.0)
            g_total = float(inv.get("grand_total") or 0.0)

            ws1.cell(row=cur_row, column=1, value=idx).alignment = align_center
            ws1.cell(row=cur_row, column=2, value=inv.get("invoice_number", "N/A")).alignment = align_center
            ws1.cell(row=cur_row, column=3, value=dt_clean).alignment = align_center
            ws1.cell(row=cur_row, column=4, value=c_name).alignment = align_left
            ws1.cell(row=cur_row, column=5, value=c_phone).alignment = align_center
            ws1.cell(row=cur_row, column=6, value=c_addr).alignment = align_left
            ws1.cell(row=cur_row, column=7, value=p_mode).alignment = align_center
            ws1.cell(row=cur_row, column=8, value=len(items)).alignment = align_center
            
            c_sub = ws1.cell(row=cur_row, column=9, value=sub)
            c_sub.alignment = align_right
            c_sub.number_format = '₹#,##0.00'

            c_disc = ws1.cell(row=cur_row, column=10, value=disc)
            c_disc.alignment = align_right
            c_disc.number_format = '₹#,##0.00'

            c_gt = ws1.cell(row=cur_row, column=11, value=g_total)
            c_gt.alignment = align_right
            c_gt.number_format = '₹#,##0.00'

            is_even = (idx % 2 == 0)
            for c in range(1, 12):
                cell = ws1.cell(row=cur_row, column=c)
                cell.font = font_data
                cell.border = border_cell
                if is_even:
                    cell.fill = fill_zebra

            ws1.row_dimensions[cur_row].height = 20
            cur_row += 1

        # Summary / Totals Row
        if len(invoices) > 0:
            ws1.cell(row=cur_row, column=1, value="").border = border_total
            ws1.cell(row=cur_row, column=2, value="TOTALS").font = font_total
            ws1.cell(row=cur_row, column=2).alignment = align_center
            ws1.cell(row=cur_row, column=2).border = border_total

            for c in range(3, 8):
                ws1.cell(row=cur_row, column=c, value="").border = border_total

            # Total Items Sum
            c_items_sum = ws1.cell(row=cur_row, column=8, value=f"=SUM(H{start_data_row}:H{cur_row-1})")
            c_items_sum.font = font_total
            c_items_sum.alignment = align_center
            c_items_sum.border = border_total

            # Total Subtotal Sum
            c_sub_sum = ws1.cell(row=cur_row, column=9, value=f"=SUM(I{start_data_row}:I{cur_row-1})")
            c_sub_sum.font = font_total
            c_sub_sum.alignment = align_right
            c_sub_sum.number_format = '₹#,##0.00'
            c_sub_sum.border = border_total

            # Total Discount Sum
            c_disc_sum = ws1.cell(row=cur_row, column=10, value=f"=SUM(J{start_data_row}:J{cur_row-1})")
            c_disc_sum.font = font_total
            c_disc_sum.alignment = align_right
            c_disc_sum.number_format = '₹#,##0.00'
            c_disc_sum.border = border_total

            # Grand Total Sum
            c_gt_sum = ws1.cell(row=cur_row, column=11, value=f"=SUM(K{start_data_row}:K{cur_row-1})")
            c_gt_sum.font = font_total
            c_gt_sum.alignment = align_right
            c_gt_sum.number_format = '₹#,##0.00'
            c_gt_sum.border = border_total

            for c in range(1, 12):
                ws1.cell(row=cur_row, column=c).fill = fill_total
            ws1.row_dimensions[cur_row].height = 24

        # ---------------- SHEET 2: ITEM-LEVEL BREAKDOWN ----------------
        ws2 = wb.create_sheet(title="Item-Level Sales Details")
        ws2.views.sheetView[0].showGridLines = True

        ws2.merge_cells("A1:K1")
        ws2["A1"] = "RC MOBILES & SERVICES - DETAILED ITEM-WISE SALES REGISTER"
        ws2["A1"].font = font_title

        headers_s2 = [
            "SL #", "Invoice Number", "Date", "Customer Name", 
            "Product / Item Name", "Brand", "HSN Code", "IMEI / Serial",
            "Qty", "Rate (₹)", "Line Total (₹)"
        ]

        row_idx_2 = 3
        for col_idx, header_text in enumerate(headers_s2, 1):
            cell = ws2.cell(row=row_idx_2, column=col_idx, value=header_text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_cell
        ws2.row_dimensions[row_idx_2].height = 26

        cur_item_row = 4
        item_sl = 1

        for inv in invoices:
            inv_num = inv.get("invoice_number", "N/A")
            dt_raw = inv.get("invoice_date") or inv.get("created_at") or ""
            dt_clean = dt_raw.split(".")[0].replace("T", " ") if dt_raw else "N/A"
            c_name = (inv.get("customer_name") or "Cash Customer").strip()

            for itm in (inv.get("invoice_items") or []):
                p_name = itm.get("product_name") or itm.get("name") or "Product"
                p_brand = itm.get("brand") or "-"
                hsn = itm.get("hsn_code") or "8517"
                imei = itm.get("imei") or itm.get("serial_number") or "-"
                qty = int(itm.get("quantity") or itm.get("qty") or 1)
                rate = float(itm.get("unit_price") or itm.get("price") or 0.0)
                tot = float(itm.get("total") or (qty * rate))

                ws2.cell(row=cur_item_row, column=1, value=item_sl).alignment = align_center
                ws2.cell(row=cur_item_row, column=2, value=inv_num).alignment = align_center
                ws2.cell(row=cur_item_row, column=3, value=dt_clean).alignment = align_center
                ws2.cell(row=cur_item_row, column=4, value=c_name).alignment = align_left
                ws2.cell(row=cur_item_row, column=5, value=p_name).alignment = align_left
                ws2.cell(row=cur_item_row, column=6, value=p_brand).alignment = align_center
                ws2.cell(row=cur_item_row, column=7, value=hsn).alignment = align_center
                ws2.cell(row=cur_item_row, column=8, value=imei).alignment = align_center
                ws2.cell(row=cur_item_row, column=9, value=qty).alignment = align_center
                
                c_r = ws2.cell(row=cur_item_row, column=10, value=rate)
                c_r.alignment = align_right
                c_r.number_format = '₹#,##0.00'

                c_t = ws2.cell(row=cur_item_row, column=11, value=tot)
                c_t.alignment = align_right
                c_t.number_format = '₹#,##0.00'

                is_even = (item_sl % 2 == 0)
                for c in range(1, 12):
                    cell = ws2.cell(row=cur_item_row, column=c)
                    cell.font = font_data
                    cell.border = border_cell
                    if is_even:
                        cell.fill = fill_zebra

                ws2.row_dimensions[cur_item_row].height = 20
                cur_item_row += 1
                item_sl += 1

        # Auto-adjust column widths for both sheets
        for sheet in [ws1, ws2]:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        lines = str(cell.value).split("\n")
                        for line in lines:
                            max_len = max(max_len, len(line))
                sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"RC_Mobiles_Sales_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return output.getvalue(), 200, {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": f"attachment; filename={filename}"
        }

    except Exception as err:
        print("Excel export error:", err)
        return jsonify({"error": str(err)}), 500

if __name__ == "__main__":
    print("==================================================")
    print("   RC MOBILES - FLASK BACKEND SERVER LAUNCHED")
    print(f"   Supabase URL: {SUPABASE_URL}")
    print("   Listening on: http://localhost:5000")
    print("==================================================")
    app.run(host="0.0.0.0", port=5000, debug=True)
